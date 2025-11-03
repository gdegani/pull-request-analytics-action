#!/usr/bin/env python3
"""Convert report/report.tsv to report/report.xlsx with proper types.

- date_iso and date_iso_end -> Excel dates (YYYY-MM-DD)
- *_hours -> numeric (float hours)
"""
import csv
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except Exception:
    # openpyxl is optional in CI/local; this archived copy is retained for history
    Workbook = None
    Font = None
    def get_column_letter(i):
        return i
import os


def parse_date(value: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def main():
    # if TSV doesn't exist, nothing to do
    tsv_path = "report/report.tsv"
    if not os.path.exists(tsv_path):
        # nothing to convert
        print(f"TSV file not found at {tsv_path}, skipping XLSX conversion.")
        return 0
    if Workbook is None:
        print("openpyxl not available; skipping conversion in archived script.")
        return 0
    wb = Workbook()
    ws = wb.active

    with open("report/report.tsv", "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        try:
            headers = next(reader)
        except StopIteration:
            headers = []

        # write header row
        ws.append(headers)
        bold = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = bold

        # determine special columns
        try:
            date_iso_idx = headers.index("date_iso")
        except ValueError:
            date_iso_idx = None
        try:
            date_iso_end_idx = headers.index("date_iso_end")
        except ValueError:
            date_iso_end_idx = None

        hour_idxs = [i for i, h in enumerate(headers) if h.endswith("_hours")]

        for row in reader:
            out = []
            for i, v in enumerate(row):
                if v == "":
                    out.append(None)
                    continue

                if date_iso_idx is not None and i == date_iso_idx:
                    parsed = parse_date(v)
                    out.append(parsed if parsed is not None else v)
                elif date_iso_end_idx is not None and i == date_iso_end_idx:
                    parsed = parse_date(v)
                    out.append(parsed if parsed is not None else v)
                elif i in hour_idxs:
                    try:
                        out.append(float(v))
                    except Exception:
                        out.append(None)
                else:
                    out.append(v)

            ws.append(out)

        # set number formats for date and hour columns
        if date_iso_idx is not None:
            col = get_column_letter(date_iso_idx + 1)
            for r in range(2, ws.max_row + 1):
                cell = ws[f"{col}{r}"]
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

        if date_iso_end_idx is not None:
            col = get_column_letter(date_iso_end_idx + 1)
            for r in range(2, ws.max_row + 1):
                cell = ws[f"{col}{r}"]
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

        for idx in hour_idxs:
            col = get_column_letter(idx + 1)
            for r in range(2, ws.max_row + 1):
                cell = ws[f"{col}{r}"]
                if cell.value is not None:
                    cell.number_format = "0.00"

    wb.save("report/report.xlsx")
    return 0


if __name__ == "__main__":
    main()
